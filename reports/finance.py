from datetime import date, datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


GRAY_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center")


# -------------------------------------------------------


def month_range(year: int, month: int):
    start = date(year, month, 1)

    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)

    return start, end


def overlap_nights(a_start, a_end, b_start, b_end):
    start = max(a_start, b_start)
    end = min(a_end, b_end)

    if start >= end:
        return 0

    return (end - start).days


# -------------------------------------------------------


def build_finance_report(rows):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансы"

    headers = [
        "Месяц",
        "Помещение",

        "Ночей в месяце",

        "Реализовано €",
        "Потенциал €",

        "Нереализовано €",

        "Расходы €",

        "Чистая прибыль €",

        "Загрузка %",
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.border = GRAY_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 24

    ws.row_dimensions[1].height = 26

    # --------------------------------------------------
    # группировка: (year, month) -> flat -> data
    # --------------------------------------------------

    grouped = defaultdict(lambda: defaultdict(lambda: {
        "nights": 0,
        "realized": 0,
        "unrealized": 0,
        "potential": 0,
        "expenses": 0,
        "price_seen": set(),
    }))

    today = date.today()

    for r in rows:

        try:
            start = datetime.fromisoformat(r["start_date"]).date()
            end = datetime.fromisoformat(r["end_date"]).date()

            price = int(r["price_per_day"])
            flat = r["flat_number"]
            fixed = float(r.get("fixed_per_booking") or 0)


        except Exception:
            continue

        cur = start.replace(day=1)

        while cur < end:

            m_start, m_end = month_range(cur.year, cur.month)
            days_in_month = (m_end - m_start).days

            nights = overlap_nights(start, end, m_start, m_end)

            if nights:

                bucket = grouped[(cur.year, cur.month)][flat]

                bucket["nights"] += nights

                # ---- реализовано ----
                lived = overlap_nights(
                    start,
                    min(today, end),
                    m_start,
                    m_end,
                )

                realized = lived * price
                bucket["realized"] += realized

                # ---- нереализовано внутри договора ----
                unrealized = (nights - lived) * price
                bucket["unrealized"] += unrealized

                # ---- потенциал квартиры за месяц ----
                # берем максимум по тарифу, если в месяце были разные договоры
                bucket["price_seen"].add(price)
                bucket["potential"] = days_in_month * max(bucket["price_seen"])

                # ---- расходы ТОЛЬКО в месяце заезда ----
                if cur.year == start.year and cur.month == start.month:
                    bucket["expenses"] += fixed

            # следующий месяц
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

    # --------------------------------------------------
    # вывод
    # --------------------------------------------------

    for (year, month) in sorted(grouped.keys(), reverse=True):

        ws.append([])
        ws.append([f"{month:02}.{year}"])
        ws.merge_cells(
            start_row=ws.max_row,
            start_column=1,
            end_row=ws.max_row,
            end_column=len(headers),
        )

        title = ws.cell(ws.max_row, 1)
        title.font = Font(bold=True)
        title.alignment = CENTER
        title.border = GRAY_BORDER

        month_realized = 0
        month_unrealized = 0
        month_potential = 0
        month_expenses = 0
        month_profit = 0
        month_nights = 0

        m_start, m_end = month_range(year, month)
        days_in_month = (m_end - m_start).days

        for flat in sorted(grouped[(year, month)]):

            b = grouped[(year, month)][flat]

            load = round(b["nights"] / days_in_month * 100, 1)

            profit = round(b["realized"] - b["expenses"], 2)

            month_realized += b["realized"]
            month_unrealized += b["unrealized"]
            month_potential += b["potential"]
            month_expenses += b["expenses"]
            month_profit += profit
            month_nights += b["nights"]

            ws.append([
                f"{month:02}.{year}",
                flat,

                b["nights"],

                f"{b['realized']} €",
                f"{b['potential']} €",

                f"{b['unrealized']} €",

                f"{b['expenses']} €",

                f"{profit} €",

                f"{load} %",
            ])

        # ---- итог месяца ----

        flats_count = len(grouped[(year, month)])
        total_possible_nights = flats_count * days_in_month

        month_load = round(
            month_nights / total_possible_nights * 100,
            1,
        ) if total_possible_nights else 0

        ws.append([])
        
        month_realized = round(month_realized, 2)
        month_unrealized = round(month_unrealized, 2)
        month_potential = round(month_potential, 2)
        month_expenses = round(month_expenses, 2)
        month_profit = round(month_profit, 2)

        ws.append([
            "ИТОГО",
            f"{flats_count} квартир",

            month_nights,

            f"{month_realized} €",
            f"{month_potential} €",

            f"{month_unrealized} €",

            f"{month_expenses} €",

            f"{month_profit} €",

            f"{month_load} %",
        ])

    # --------------------------------------------------
    # форматирование всех ячеек
    # --------------------------------------------------

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

        for cell in row:
            if cell.value is not None:
                cell.alignment = CENTER
                cell.border = GRAY_BORDER

    path = "/tmp/financial_report.xlsx"
    wb.save(path)
    return path
