from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def build_expenses_report(rows):

    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"

    CENTER = Alignment(horizontal="center")
    BOLD = Font(bold=True)

    # -------------------------
    # СВОДКА
    # -------------------------

    total_all = 0.0
    total_cash = 0.0
    total_company = 0.0

    for r in rows:
        amt = float(r["amount"])
        total_all += amt

        if r["payment_method"] == "cash":
            total_cash += amt
        else:
            total_company += amt

    ws["A1"] = "СВОДКА"
    ws["A1"].font = BOLD

    ws.append(["Всего потрачено €", round(total_all, 2)])
    ws.append(["Наличными €", round(total_cash, 2)])
    ws.append(["С фирмы €", round(total_company, 2)])

    ws.append([])
    ws.append([])

    # -------------------------
    # ГРУППИРОВКА ПО МЕСЯЦАМ
    # -------------------------

    grouped = defaultdict(list)

    for r in rows:
        dt = datetime.fromisoformat(r["expense_date"])
        key = f"{dt.year}-{dt.month:02}"
        grouped[key].append(r)

    # -------------------------
    # ВЫВОД ПО МЕСЯЦАМ
    # -------------------------

    for ym in sorted(grouped.keys(), reverse=True):

        year, month = ym.split("-")

        ws.append([f"{month}.{year}"])
        ws.cell(ws.max_row, 1).font = BOLD

        ws.append([
            "Дата",
            "Описание",
            "Сумма €",
            "Способ оплаты",
        ])

        ws.row_dimensions[ws.max_row].height = 22

        m_total = 0.0
        m_cash = 0.0
        m_company = 0.0

        for r in grouped[ym]:

            amt = float(r["amount"])
            m_total += amt

            if r["payment_method"] == "cash":
                m_cash += amt
            else:
                m_company += amt

            dt = datetime.fromisoformat(r["expense_date"]).strftime("%d.%m.%Y")

            pay = "Наличные" if r["payment_method"] == "cash" else "С фирмы"

            desc = r.get("description") or r.get("comment") or ""

            ws.append([
                dt,
                desc,
                round(amt, 2),
                pay,
            ])

        ws.append([])
        ws.append([
            "ИТОГО:",
            "",
            round(m_total, 2),
            "",
        ])

        ws.append([
            "Наличными:",
            "",
            round(m_cash, 2),
            "",
        ])

        ws.append([
            "С фирмы:",
            "",
            round(m_company, 2),
            "",
        ])

        ws.append([])
        ws.append([])

    # -------------------------
    # ширина колонок
    # -------------------------

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 30

    path = "/tmp/expenses_report.xlsx"
    wb.save(path)

    return path
