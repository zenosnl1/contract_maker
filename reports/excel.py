from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from core.constants import STAT_COLUMNS, STATS_HEADERS


def build_stats_excel(rows):

    # сортировка от новых к старым по дате заезда
    rows = sorted(
        rows,
        key=lambda r: r.get("start_date") or "",
        reverse=True,
    )

    wb = Workbook()

    gray_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    center_align = Alignment(horizontal="center", vertical="center")

    # ====== СВОДКА ======

    ws1 = wb.active
    ws1.title = "Сводка"

    total_income = sum(int(r.get("total_price") or 0) for r in rows)
    total_nights = sum(int(r.get("nights") or 0) for r in rows)

    first_date = min(
        r.get("start_date") for r in rows
        if r.get("start_date")
    )

    ws1.append(["Общий доход (€)", total_income])
    ws1.append(["Всего ночей", total_nights])
    ws1.append(["Дата первого договора", first_date])

    for row in ws1.iter_rows():
        ws1.row_dimensions[row[0].row].height = 20

        for cell in row:
            cell.font = Font(bold=cell.column == 1)
            cell.alignment = center_align
            cell.border = gray_border

            ws1.column_dimensions[get_column_letter(cell.column)].width = 30

    # ====== ДОГОВОРЫ ======

    ws2 = wb.create_sheet("Договоры")

    if not rows:
        return None

    # порядок в начале
    preferred = [
        "contract_code",
        "flat_number",
        "client_name",
        "client_id",
        "client_number",

        "start_date",
        "end_date",
        "actual_checkout_date",

        "nights",
        "price_per_day",
        "total_price",

        "deposit",
        "returned_deposit",
        "deposit_comment",

        "payment_method",
        "invoice_issued",
        "invoice_number",

        "is_closed",
    ]

    # все реально пришедшие поля
    all_keys = list(rows[0].keys())

    # ❌ технические поля исключаем
    banned = {"id", "created_at"}

    all_keys = [k for k in all_keys if k not in banned]

    # объединяем: сначала preferred, потом остальные
    keys = preferred + [k for k in all_keys if k not in preferred]

    headers_map = {
        "contract_code": "Код договора",
        "flat_number": "Помещение",

        "client_name": "Имя клиента",
        "client_id": "Документ",
        "client_address": "Адрес",
        "client_mail": "Email",
        "client_number": "Телефон",

        "start_date": "Дата заезда",
        "end_date": "Дата выезда",
        "actual_checkout_date": "Фактический выезд",

        "nights": "Ночей",
        "price_per_day": "Цена / ночь",
        "total_price": "Общая сумма",

        "deposit": "Депозит",
        "returned_deposit": "Возвращено",
        "deposit_comment": "Комментарий",

        "payment_method": "Способ оплаты",
        "invoice_issued": "Счёт выставлен",
        "invoice_number": "Номер счёта",

        "is_closed": "Статус договора",
    }

    # ---- Заголовки ----

    ws2.append([headers_map.get(k, k) for k in keys])

    for col in range(1, len(keys) + 1):

        cell = ws2.cell(row=1, column=col)

        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = gray_border

        ws2.column_dimensions[get_column_letter(col)].width = 26

    ws2.row_dimensions[1].height = 26

    # ---- Данные ----

    def humanize_value(key, val):

        if key == "is_closed":
            return "Завершён" if val else "Активен"

        if key == "payment_method":
            return {
                "cash": "Наличные",
                "bank_transfer": "Банковский перевод",
            }.get(val, "-")

        if key == "invoice_issued":
            return "Да" if val else "Нет"

        return val

    for r in rows:
        ws2.append([humanize_value(k, r.get(k)) for k in keys])

    for row in ws2.iter_rows(min_row=2):

        ws2.row_dimensions[row[0].row].height = 18

        for cell in row:
            cell.alignment = center_align
            cell.border = gray_border

    path = "/tmp/contracts_stats.xlsx"
    wb.save(path)

    return path
